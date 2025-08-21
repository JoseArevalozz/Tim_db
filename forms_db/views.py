import csv
import pandas as pd
import xlwt
import xlsxwriter
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from plotly.offline import plot
from io import StringIO
from io import BytesIO
from forms_db.module import WriteToExcel
from datetime import date, datetime, timedelta, time
from django.views.decorators.http import require_GET
from django.db import transaction, models
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.contrib.auth.models import User
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q, F
from django.shortcuts import get_object_or_404, redirect
from django.utils import timezone
from django.core.paginator import Paginator
from .forms import EmployeesForm, UutForm, FailureForm, BoomForm, RejectedForm, ErrorMessageForm, StationForm, MaintenanceForm, SpareForm, ReleaseForm, CorrectiveMaintenanceForm, ManualFailureRegistrationForm
from .models import Uut, Employes, Failures, Station, ErrorMessages, Booms, Rejected, Release, Maintenance, SparePart, TestHistory  
from django.views.decorators.csrf import csrf_exempt

@login_required(login_url='login')
def home(request):
    employe = Employes.objects.get(employeeNumber=request.user)

    # Determinar si es supervisor
    is_supervisor = employe.privileges.endswith('S')
    base_project = employe.privileges[:-1] if is_supervisor else employe.privileges
    
    # Obtener la fecha actual
    today = timezone.now().date()
    
    # Obtener estadísticas de pruebas para el día actual
    test_history = TestHistory.objects.filter(
        uut__pn_b__project=base_project,
        test_date__date=today
    )
    
    # Calcular métricas
    total_tests = test_history.count()
    passed_tests = test_history.filter(status=True).count()
    
    # Calcular yield (con protección contra división por cero)
    today_yield = round((passed_tests / total_tests * 100), 2) if total_tests > 0 else 0
    
    active_failures_count = Uut.objects.filter(status=True, pn_b__project=base_project).count()
    pending_rejects_count = Failures.objects.filter(status=True, sn_f__pn_b__project=base_project).count()
    
    context = {
        'employe': employe,
        'active_failures_count': active_failures_count,
        'pending_rejects_count': pending_rejects_count,
        'today_yield': today_yield,
        'total_tests_today': total_tests,
        'passed_tests_today': passed_tests,
        'is_supervisor': is_supervisor,
        'base_project': base_project
    }
    
    if 'bt-project' in request.POST: 
        if request.method == 'POST':
            employe.privileges = request.POST.get('bt-project')
            employe.save()
            return redirect('home')

    return render(request=request, template_name='base/first.html', context=context)

def loginUser(request):
    uuts = Uut.objects.all()
    
    if request.user.is_authenticated:
        return redirect('home')
    
    if request.method == 'POST':
        username = request.POST.get('user').lower()
        password = request.POST.get('password')
        
        try:
            user = User.objects.get(username=username)
        except:
            messages.error(request=request, message='User does not exist')
            
        user = authenticate(request=request, username=username, password=password)
        
        if user is not None:
            login(request=request, user=user)
            return redirect('home')
        else:
            messages.error(request=request, message='Username or password does not exist')
        
    context={'uuts':uuts}
    return render(request=request, template_name='base/first.html', context=context)


def logoutUser(request):
    logout(request=request)
    return redirect('home')

def passwordForm(request):
    u = User.objects.get(username__exact=request.user)

    if request.method == 'POST':
        if request.POST.get('new-password') == request.POST.get('val-password'):
            u.set_password(request.POST.get('new-password'))
            u.save()
            return redirect('home')
        else:
            messages.error(request=request, message='The new password its not the same')
    context = {}
    return render(request=request, template_name='base/password_form.html', context=context)

@login_required(login_url='login')
def employeesForm(request):
    employe = Employes.objects.get(employeeNumber=request.user)
    is_supervisor = employe.privileges.endswith('S')
    base_project = employe.privileges[:-1] if is_supervisor else employe.privileges
    
    form = EmployeesForm()
    if 'bt-project' in request.POST:
        if request.method == 'POST':
            employe.privileges = request.POST.get('bt-project')
            employe.save()
            return redirect('employees_form')
    
    if base_project == 'NA':
        return redirect('home')
    
    if request.method == 'POST':
        form = EmployeesForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('home')
    
    context = {'form': form, 'employe': employe, 'is_supervisor': is_supervisor, 'base_project': base_project}
    return render(request=request, template_name='base/employee_form.html', context=context)

@login_required(login_url='login')
def uutForm(request):
    employe = Employes.objects.get(employeeNumber=request.user)
    is_supervisor = employe.privileges.endswith('S')
    base_project = employe.privileges[:-1] if is_supervisor else employe.privileges
    
    form = UutForm()
    form.fields['pn_b'].queryset = Booms.objects.filter(project=base_project)
    
    if 'bt-project' in request.POST: 
        if request.method == 'POST':
            employe.privileges = request.POST.get('bt-project')
            employe.save()
            return redirect('uut_form')
    
    if base_project == 'NA':
        return redirect('home')
    
    if base_project == 'SONY':
        form.fields['pn_b'].queryset = Booms.objects.filter(Q(commodity='RACK') | Q(commodity='SLED') | Q(commodity='KURA'))
        
    if request.method == 'POST':
        pn_booms = Booms.objects.get(pn=request.POST.get('pn_b'))
        try:
            Uut.objects.create(
                sn=request.POST.get('sn'),
                pn_b=pn_booms,
                employee_e=employe,
                status = True if request.POST.get('status') == 'on' else False,
            )
            return redirect('showUuts')
        except:
            messages.error(request=request, message='PN already registered!')
    
    context = {'form':form, 'employe': employe, 'is_supervisor': is_supervisor, 'base_project': base_project}
    return render(request=request, template_name='base/uut_form.html', context=context)

@login_required(login_url='login')
def failureForm(request, pk):
    employe = Employes.objects.get(employeeNumber=request.user)
    is_supervisor = employe.privileges.endswith('S')
    base_project = employe.privileges[:-1] if is_supervisor else employe.privileges
    
    if not Uut.objects.filter(sn=pk).exists():
        return redirect('showUuts')
    
    uut = Uut.objects.get(sn=pk)
    last_failure = Failures.objects.filter(sn_f=uut).order_by('-failureDate').first()
    
    if not last_failure:
        return redirect('showUuts')

    if request.method == 'POST':
        if 'bt-project' in request.POST:
            employe.privileges = request.POST.get('bt-project')
            employe.save()
            return redirect('showUuts')
        
        if base_project == 'NA':
            return redirect('home')
        
        form = FailureForm(request.POST, request.FILES)
        
        if form.is_valid():
            failure = last_failure
            failure.analysis = form.cleaned_data['analysis']
            failure.rootCauseCategory = form.cleaned_data['rootCauseCategory']
            failure.defectSymptom = form.cleaned_data['defectSymptom']
            failure.correctiveActions = form.cleaned_data['correctiveActions']
            failure.comments = form.cleaned_data['comments']
            failure.status = form.cleaned_data['status']
            failure.employee_e = employe
            
            if 'imgEvindence' in request.FILES:
                failure.imgEvindence = form.cleaned_data['imgEvindence']
            
            failure.save()
            return redirect('showUuts')
    else:
        initial_data = {
            'analysis': last_failure.analysis,
            'rootCauseCategory': last_failure.rootCauseCategory,
            'defectSymptom': last_failure.defectSymptom,
            'correctiveActions': last_failure.correctiveActions,
            'comments': last_failure.comments,
            'error_message': str(last_failure.id_er) if last_failure.id_er else '',
            'status': False  # Valor por defecto
        }
        form = FailureForm(initial=initial_data, instance=last_failure)
    
    context = {
        'form': form,
        'employe': employe,
        'uut': uut,
        'is_supervisor': is_supervisor,
        'base_project': base_project
    }
    return render(request, 'base/failure_form.html', context)

@login_required(login_url='login')
def menu_pruebas(request):
    employe = Employes.objects.get(employeeNumber=request.user)
    is_supervisor = employe.privileges.endswith('S')
    base_project = employe.privileges[:-1] if is_supervisor else employe.privileges
    
    if 'bt-project' in request.POST: 
        if request.method == 'POST':
            employe.privileges = request.POST.get('bt-project')
            employe.save()
        return redirect('menu_pruebas')
    
    if base_project == 'NA':
        return redirect('home')
    
    context = {'employe': employe, 'is_supervisor': is_supervisor, 'base_project': base_project}
    return render(request=request, template_name='base/menuPruebas.html', context=context)

@login_required(login_url='login')
def menu_registros(request):
    employe = Employes.objects.get(employeeNumber=request.user)
    is_supervisor = employe.privileges.endswith('S')
    base_project = employe.privileges[:-1] if is_supervisor else employe.privileges
    
    if 'bt-project' in request.POST: 
        if request.method == 'POST':
            employe.privileges = request.POST.get('bt-project')
            employe.save()
        return redirect('menu_registros')
    
    if base_project == 'NA':
        return redirect('home')
    
    context={'employe': employe, 'is_supervisor': is_supervisor, 'base_project': base_project}
    return render(request=request, template_name='base/menuRegistros.html', context=context)

@login_required(login_url='login')
def menu_metricas(request):
    employe = Employes.objects.get(employeeNumber=request.user)
    is_supervisor = employe.privileges.endswith('S')
    base_project = employe.privileges[:-1] if is_supervisor else employe.privileges
    
    if 'bt-project' in request.POST: 
        if request.method == 'POST':
            employe.privileges = request.POST.get('bt-project')
            employe.save()
        return redirect('menu_metricas')
    
    if base_project == 'NA':
        return redirect('home')
    
    context = {'employe':employe, 'is_supervisor': is_supervisor, 'base_project': base_project}
    return render(request=request, template_name='base/menuMetric.html', context = context)

@login_required(login_url='login')
def showUuts(request):
    employe = Employes.objects.get(employeeNumber=request.user)
    is_supervisor = employe.privileges.endswith('S')
    base_project = employe.privileges[:-1] if is_supervisor else employe.privileges
    
    q = request.GET.get('q') if request.GET.get('q') != None else ''
    
    uuts = Uut.objects.filter(status=True).filter(pn_b__project=base_project).filter(
        Q(sn__icontains=q) |
        Q(pn_b__pn__icontains=q) |
        Q(date__icontains=q)
    ).order_by('-date')

    if 'bt-project' in request.POST: 
        if request.method == 'POST':
            employe.privileges = request.POST.get('bt-project')
            employe.save()
            return redirect('showUuts')
            
    if base_project == 'NA':
        return redirect('home')
    
    context = {
        'uuts': uuts, 
        'employe': employe, 
        'search_bt': True,
        'is_supervisor': is_supervisor,
        'base_project': base_project
    }
    
    return render(request=request, template_name='base/showUuts.html', context=context)

@login_required(login_url='login')
def boomForm(request):
    employe = Employes.objects.get(employeeNumber=request.user)
    is_supervisor = employe.privileges.endswith('S')
    base_project = employe.privileges[:-1] if is_supervisor else employe.privileges
    
    form = BoomForm()
    
    if base_project == 'DELL':
        list_products = [('Senna','Senna'), ('Pathfinder','Pathfinder'), ('Sojouner','Sojouner'), ('Hook','Hook'), ('Outlander','Outlander'), ('Minerrall Well','Minerrall Well'), ('MMCs','MCCs'), ('Fornax SAM','Fornax SAM'), ('Fornax DIB','Fornax DIB'), ('Fornax CIT','Fornax CIT'), ('Indus DIB','Indus DIB'), ('Indus BOP','Indus BOP'), ('Indus SAM','Indus SAM'), ('Indus CIT','Indus CIT')]
        form.fields['product'].widget.choices = list_products
        
    if base_project == 'PMDU':
        list_products = [('PMDU','PMDU')]
        form.fields['product'].widget.choices = list_products
        
    if base_project == '1G-SW':
        list_products = [('Switch','Switch')]
        
    if base_project == 'SONY':
        list_products = [('CRONOS', 'CRONOS')]
        form.fields['product'].widget.choices = list_products
        
    if 'bt-project' in request.POST: 
        if request.method == 'POST':
            employe.privileges = request.POST.get('bt-project')
            employe.save()
            return redirect('boom_form')
    
    if base_project == 'NA':
        return redirect('home')
    
    if request.method == 'POST':
        try: 
            Booms.objects.create(
                pn=request.POST.get('pn'),
                description=request.POST.get('description'),
                commodity=request.POST.get('commodity'),
                product=request.POST.get('product'),
                ubiLogic=request.POST.get('ubiLogic'),
                employee_e=employe,
                project=base_project
            )
            return redirect('home')
        except:
            messages.error(request=request, message='PN already registered!')
            
    context = {
        'form': form, 
        'employe': employe,
        'is_supervisor': is_supervisor,
        'base_project': base_project
    }
    return render(request=request, template_name='base/boom_form.html', context=context)

@login_required(login_url='login')
def rejectedForm(request, pk):
    employe = Employes.objects.get(employeeNumber=request.user)
    is_supervisor = employe.privileges.endswith('S')
    base_project = employe.privileges[:-1] if is_supervisor else employe.privileges
    
    form = RejectedForm()
    form.fields['pn_b'].queryset = Booms.objects.filter(project=base_project)
    
    if 'bt-project' in request.POST: 
        if request.method == 'POST':
            employe.privileges = request.POST.get('bt-project')
            employe.save()
            return redirect('showRejecteds')
    
    if base_project == 'NA':
        return redirect('home')
       
    if request.method == 'POST':
        failure = Failures.objects.get(id=pk)
        pn_booms = Booms.objects.get(pn=request.POST.get('pn_b'))
        
        Rejected.objects.create(
            id_f=failure,
            pn_b=pn_booms,
            snDamaged=request.POST.get('snDamaged'),
            snNew=request.POST.get('snNew'),
            folio=request.POST.get('folio'),
            employee_e=employe
        )
        
        Failures.objects.filter(id=pk).update(status=False)
        return redirect('showRejecteds')
    
    context = {
        'form': form, 
        'employe': employe,
        'is_supervisor': is_supervisor,
        'base_project': base_project
    }
    return render(request=request, template_name='base/rejected_form.html', context=context)

@login_required(login_url='login')
def showRejecteds(request):
    employe = Employes.objects.get(employeeNumber=request.user)
    is_supervisor = employe.privileges.endswith('S')
    base_project = employe.privileges[:-1] if is_supervisor else employe.privileges
    
    q = request.GET.get('q') if request.GET.get('q') != None else ''
    
    failures = Failures.objects.filter(status=True).filter(sn_f__pn_b__project=base_project).filter(
        Q(sn_f__sn__icontains=q)
    ).order_by('-failureDate')
    
    if 'bt-project' in request.POST: 
        if request.method == 'POST':
            employe.privileges = request.POST.get('bt-project')
            employe.save()
            return redirect('showRejecteds')
    
    if base_project == 'NA':
        return redirect('home')
    
    context = {
        'failures': failures, 
        'employe': employe, 
        'search_bt': True,
        'is_supervisor': is_supervisor,
        'base_project': base_project
    }
    
    return render(request=request, template_name='base/showRejected.html', context=context)

@login_required(login_url='login')
def errorMessageForm(request):
    employe = Employes.objects.get(employeeNumber=request.user)
    is_supervisor = employe.privileges.endswith('S')
    base_project = employe.privileges[:-1] if is_supervisor else employe.privileges
    
    form = ErrorMessageForm()
    form.fields['pn_b'].queryset = Booms.objects.filter(project=base_project)
    
    if 'bt-project' in request.POST: 
        if request.method == 'POST':
            employe.privileges = request.POST.get('bt-project')
            employe.save()
            return redirect('errorMessage_form')
    
    if base_project == 'NA':
        return redirect('home')
    
    if base_project == 'SONY':
        form.fields['pn_b'].queryset = Booms.objects.filter(Q(commodity='RACK') | Q(commodity='SLED') | Q(commodity='KURA'))
    
    if request.method == 'POST':
        pn_booms = Booms.objects.get(pn=request.POST.get('pn_b'))
        
        ErrorMessages.objects.create(
            message=request.POST.get('message'),
            employee_e=employe,
            pn_b=pn_booms
        )
        return redirect('errorMessage_form')
    
    context = {
        'form': form, 
        'employe': employe,
        'is_supervisor': is_supervisor,
        'base_project': base_project
    }
    return render(request=request, template_name='base/errorMessage.html', context=context)

@login_required(login_url='login')
def stationForm(request):
    employe = Employes.objects.get(employeeNumber=request.user)
    is_supervisor = employe.privileges.endswith('S')
    base_project = employe.privileges[:-1] if is_supervisor else employe.privileges
    
    form = StationForm()
    if 'bt-project' in request.POST: 
        if request.method == 'POST':
            employe.privileges = request.POST.get('bt-project')
            employe.save()
            return redirect('station_form')
    
    if base_project == 'NA':
        return redirect('home')
    
    if request.method == 'POST':
        form = StationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('home')
    
    context = {
        'form': form, 
        'employe': employe,
        'is_supervisor': is_supervisor,
        'base_project': base_project
    }
    return render(request=request, template_name='base/station_form.html', context=context)

@login_required(login_url='login')
def correctiveMaintenanceForm(request, pn_sp, maintenance_id):
    employe = Employes.objects.get(employeeNumber=request.user)
    is_supervisor = employe.privileges.endswith('S')
    base_project = employe.privileges[:-1] if is_supervisor else employe.privileges

    if base_project == 'NA':
        return redirect('home')

    spare_part = get_object_or_404(SparePart, pn=pn_sp)

    if request.method == 'POST':
        form = CorrectiveMaintenanceForm(request.POST)
        if form.is_valid():
            maintenance = Maintenance.objects.get(id=maintenance_id)
            maintenance.dateFinish = timezone.now()
            maintenance.id_sp = spare_part

            if spare_part.quantity > 0:
                spare_part.quantity -= 1 
                spare_part.save()
            else:
                pass
                
            maintenance.status = False
            maintenance.save()
            return redirect('show_Maintenance')
    else:
        form = CorrectiveMaintenanceForm()

    context = {
        'form': form, 
        'pn_sp': pn_sp,
        'is_supervisor': is_supervisor,
        'base_project': base_project
    }
    return render(request, 'base/corrective_maintenance_form.html', context)

@login_required(login_url='login')
def maintenanceForm(request):
    employe = Employes.objects.get(employeeNumber=request.user)
    is_supervisor = employe.privileges.endswith('S')
    base_project = employe.privileges[:-1] if is_supervisor else employe.privileges
    
    if base_project == 'NA':
        return redirect('home')

    if request.method == 'POST':
        form = MaintenanceForm(request.POST)
        if form.is_valid():
            maintenance = form.save(commit=False)
            maintenance.employee_e = employe
            maintenance.dateStart = timezone.now()

            if maintenance.maintenanceType == 'Corrective':
                maintenance.dateFinish = maintenance.dateStart
                maintenance.save()
                return redirect('show_Maintenance')

            maintenance.dateFinish = maintenance.dateStart
            maintenance.status = False
            maintenance.save()
            return redirect('home')
    else:
        form = MaintenanceForm()

    form.fields['station_s'].queryset = Station.objects.filter(stationProject=base_project)
    
    context = {
        'form': form, 
        'employe': employe,
        'is_supervisor': is_supervisor,
        'base_project': base_project
    }
    return render(request, 'base/maintenance_form.html', context)

def showMaintenanceForm(request):
    employe = Employes.objects.get(employeeNumber=request.user)
    is_supervisor = employe.privileges.endswith('S')
    base_project = employe.privileges[:-1] if is_supervisor else employe.privileges

    if base_project == 'NA':
        return redirect('home')

    corrective_instances = Maintenance.objects.filter(
        maintenanceType='Corrective',
        status=True,
        station_s__stationProject=base_project
    ).select_related('station_s', 'id_sp', 'employee_e').order_by('-dateStart')

    context = {
        'corrective_instances': corrective_instances,
        'is_supervisor': is_supervisor,
        'base_project': base_project
    }
    return render(request, 'base/corrective_stations.html', context)

@login_required(login_url='login')
def spareForm(request):
    employe = Employes.objects.get(employeeNumber=request.user)
    is_supervisor = employe.privileges.endswith('S')
    base_project = employe.privileges[:-1] if is_supervisor else employe.privileges
    
    form = SpareForm()
    
    if 'bt-project' in request.POST: 
        if request.method == 'POST':
            employe.privileges = request.POST.get('bt-project')
            employe.save()
            return redirect('spare_form')
    
    if base_project == 'NA':
        return redirect('home')
    
    if request.method == 'POST':
        form = SpareForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('home')
    
    context = {
        'form': form, 
        'employe': employe,
        'is_supervisor': is_supervisor,
        'base_project': base_project
    }
    return render(request=request, template_name='base/spare_form.html', context=context)

@login_required(login_url='login')
def userPage(request, pk):
    employe = Employes.objects.get(employeeNumber=request.user)
    is_supervisor = employe.privileges.endswith('S')
    base_project = employe.privileges[:-1] if is_supervisor else employe.privileges
    
    user = Employes.objects.get(employeeNumber=pk)
    
    if 'bt-project' in request.POST: 
        if request.method == 'POST':
            employe.privileges = request.POST.get('bt-project')
            employe.save()
        return redirect('home')
    
    if base_project == 'NA':
        return redirect('home')
    
    context = {
        'user': user, 
        'employe':employe,
        'is_supervisor': is_supervisor,
        'base_project': base_project
    }
    return render(request=request, template_name='base/user.html', context=context)

@login_required(login_url='login')
def tableRejects(request):
    employe = Employes.objects.get(employeeNumber=request.user)
    is_supervisor = employe.privileges.endswith('S')
    base_project = employe.privileges[:-1] if is_supervisor else employe.privileges
    
    q = request.GET.get('q') if request.GET.get('q') != None else ''
    
    try:
        if '/' in q:
            dates = q.split('/')
            
            dateStart = list(map(int, dates[0].split('-')))
            start = date(dateStart[0], dateStart[1], dateStart[2])
            
            dateEnd = list(map(int, dates[1].split('-')))     
            end = date(dateEnd[0], dateEnd[1], dateEnd[2])
            new_end = end + timedelta(days=1)
            
            rejects = Rejected.objects.filter(id_f__sn_f__pn_b__project=base_project).filter(
                dateRejected__range=[start, new_end],).order_by('-dateRejected')
        else:
            rejects = Rejected.objects.filter(id_f__sn_f__pn_b__project=base_project).filter(
                Q(folio__icontains=q) |
                Q(id_f__sn_f__sn__icontains=q) ).order_by('-dateRejected') 
    except:
        return redirect('tableRejects')
    
    if 'bt-project' in request.POST: 
        if request.method == 'POST':
            employe.privileges = request.POST.get('bt-project')
            employe.save()
            return redirect('tableRejects')
    
    if base_project == 'NA':
        return redirect('home')
   
    if request.method == 'POST':
        check = request.POST.getlist('check')
        today = datetime.today().strftime("%Y-%m-%d_%H-%M")
        response = HttpResponse(content_type='application/vnd.ms-excel')
        file_name= f'DashB_{today}'
        response['Content-Disposition'] = f'attachment; filename={file_name}.xlsx'
        xlsx_data = WriteToExcel(check=check)
        response.write(xlsx_data)
        return response
    
    context = {
        'employe': employe, 
        'rejects': rejects, 
        'search_bt': True,
        'is_supervisor': is_supervisor,
        'base_project': base_project
    }
    return render(request=request, template_name='base/table_rejects.html', context=context)

def finish_uut(request, sn):
    # Obtener el UUT, el empleado y la última falla asociada
    uut = get_object_or_404(Uut, sn=sn)
    employee = get_object_or_404(Employes, employeeNumber=request.user)
    
    # Cambiar el estado del UUT
    uut.status = False
    uut.save()
    
    # Obtener la última falla asociada al UUT
    last_failure = Failures.objects.filter(sn_f=uut).order_by('-failureDate').first()
    
    if last_failure:
        # Actualizar SOLO el empleado que cerró el UUT
        last_failure.employee_e = employee
        last_failure.save()
    
    return redirect('showUuts')

@login_required(login_url='login')
def tableFailures(request):
    employe = Employes.objects.get(employeeNumber=request.user)
    is_supervisor = employe.privileges.endswith('S')
    base_project = employe.privileges[:-1] if is_supervisor else employe.privileges
    
    search_term = request.GET.get('q', '').strip()
    
    failures = Failures.objects.filter(
        sn_f__pn_b__project=base_project
    ).select_related(
        'sn_f', 'id_er', 'employee_e', 'id_s'
    ).order_by('-failureDate')

    if search_term:
        failures = failures.filter(
            Q(sn_f__sn__icontains=search_term) |
            Q(id_er__message__icontains=search_term) |
            Q(shiftFailure__icontains=search_term)
        )

    if request.method == 'POST':
        response = HttpResponse(content_type='application/ms-excel')
        today = datetime.now().strftime("%Y-%m-%d_%H-%M")
        response['Content-Disposition'] = f'attachment; filename="Failures_{today}.xls"'

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet("Failures")

        header_style = xlwt.XFStyle()
        header_style.font.bold = True

        columns = [
            'SN', 'Status', 'Failure Date', 'Station', 
            'Error Message', 'Analysis', 'Root Cause Category',
            'Defect Symptom', 'Employee', 'Shift',
            'Corrective Actions', 'Comments'
        ]
        
        for col_num, column in enumerate(columns):
            ws.write(0, col_num, column, header_style)

        row_style = xlwt.XFStyle()
        date_style = xlwt.XFStyle()
        date_style.num_format_str = 'YYYY-MM-DD HH:MM'

        if 'download' in request.POST:
            selected_ids = request.POST.getlist('check')
            failures_to_export = failures.filter(id__in=selected_ids)
        else:
            failures_to_export = failures

        for row_num, failure in enumerate(failures_to_export, start=1):
            ws.write(row_num, 0, failure.sn_f.sn if failure.sn_f else '', row_style)
            ws.write(row_num, 1, 'Active' if failure.status else 'Closed', row_style)
            ws.write(row_num, 2, failure.failureDate, date_style)
            ws.write(row_num, 3, failure.id_s.stationName if failure.id_s else '', row_style)
            ws.write(row_num, 4, failure.id_er.message if failure.id_er else '', row_style)
            ws.write(row_num, 5, failure.analysis or '', row_style)
            ws.write(row_num, 6, failure.rootCauseCategory or '', row_style)
            ws.write(row_num, 7, failure.defectSymptom or '', row_style)
            ws.write(row_num, 8, failure.employee_e.employeeName if failure.employee_e else '', row_style)
            ws.write(row_num, 9, failure.shiftFailure, row_style)
            ws.write(row_num, 10, failure.correctiveActions or '', row_style)
            ws.write(row_num, 11, failure.comments or '', row_style)

        wb.save(response)
        return response

    paginator = Paginator(failures, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'employe': employe,
        'failures': page_obj,
        'search_query': search_term,
        'search_bt': True,
        'is_supervisor': is_supervisor,
        'base_project': base_project
    }
    return render(request, 'base/table_fails.html', context)

@login_required(login_url='login')
def tableUuts(request):
    employe = Employes.objects.get(employeeNumber=request.user)
    is_supervisor = employe.privileges.endswith('S')
    base_project = employe.privileges[:-1] if is_supervisor else employe.privileges
    
    q = request.GET.get('q') if request.GET.get('q') != None else ''
    
    try:
        if '/' in q:
            dates = q.split('/')
            
            dateStart = list(map(int, dates[0].split('-')))
            start = date(dateStart[0], dateStart[1], dateStart[2])
            
            dateEnd = list(map(int, dates[1].split('-')))     
            end = date(dateEnd[0], dateEnd[1], dateEnd[2])
            new_end = end + timedelta(days=1)
            
            uuts = Uut.objects.filter(pn_b__project=base_project).filter(
                failureDate__range=[start, new_end],)
        else:
            uuts = Uut.objects.filter(pn_b__project=base_project).filter(
                Q(date__icontains=q)
            )
    except:
        return redirect('tableFailures')
    
    if 'bt-project' in request.POST: 
        if request.method == 'POST':
            employe.privileges = request.POST.get('bt-project')
            employe.save()
            return redirect('tableFailures')
    
    if base_project == 'NA':
        return redirect('home')
   
    if request.method == 'POST':
        check = request.POST.getlist('check')
        response = HttpResponse(content_type='application/ms-excel')
        today = datetime.today().strftime("%Y-%m-%d_%H-%M")
        response['Content-Disposition'] = f'attachment; filename="Uuts{today}.xls"'

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet("sheet1")

        row_num = 0
        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        columns = ['Sn', 'Pn', 'Model', 'Employee', 'Status']

        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], font_style)

        font_style = xlwt.XFStyle()

        for checked in check:
            uut = Uut.objects.get(sn=checked)
            
            sn = str(uut.sn)
            pn = str(uut.pn_b.pn)
            model = str(uut.pn_b.product)
            employee = str(uut.employee_e.employeeName)
            status = str(uut.status)
            
            row_num = row_num + 1
            ws.write(row_num, 0, sn, font_style)
            ws.write(row_num, 1, pn, font_style)
            ws.write(row_num, 2, model, font_style)
            ws.write(row_num, 3, employee, font_style)
            ws.write(row_num, 4, status, font_style)

        wb.save(response)
        return response
        
    context = {
        'employe': employe, 
        'uuts': uuts,
        'is_supervisor': is_supervisor,
        'base_project': base_project
    }
    return render(request=request, template_name='base/table_uuts.html', context=context)

@login_required(login_url='login')
def releaseForm(request):
    employe = Employes.objects.get(employeeNumber=request.user)
    is_supervisor = employe.privileges.endswith('S')
    base_project = employe.privileges[:-1] if is_supervisor else employe.privileges
    
    form = ReleaseForm()
    
    if 'bt-project' in request.POST: 
        if request.method == 'POST':
            employe.privileges = request.POST.get('bt-project')
            employe.save()
            return redirect('uut_form')
    
    if base_project != 'SONY':
        return redirect('home')
    
    if request.method =='POST':
        files = request.FILES
        cimsI = files.get('cims')
        crabberI = files.get('crabber')
        try:
            Release.objects.create(
                serial = request.POST.get('serial'),
                shift = request.POST.get('shift'),
                nicho = request.POST.get('nicho'),
                cims = cimsI,
                crabber = crabberI,
                employee_e = employe
            )
            return redirect('home')
        except:
            messages.error(request=request, message='SN already registered!')

    context = {
        'form':form, 
        'employe':employe,
        'is_supervisor': is_supervisor,
        'base_project': base_project
    }
    return render(request=request, template_name='base/release_form.html', context=context)

@login_required(login_url='login')
def tableRelease(request):
    employe = Employes.objects.get(employeeNumber=request.user)
    is_supervisor = employe.privileges.endswith('S')
    base_project = employe.privileges[:-1] if is_supervisor else employe.privileges
    
    start = request.GET.get('fechaI')
    end = request.GET.get('fechaF')
    
    if 'bt-project' in request.POST: 
        if request.method == 'POST':
            employe.privileges = request.POST.get('bt-project')
            employe.save()
            return redirect('uut_form')
    
    if base_project != 'SONY':
        return redirect('home')
    
    if start == None or end == None or start == '' or end == '':
        l1 = Release.objects.filter(shift='1')
        l2 = Release.objects.filter(shift='2')
        l3 = Release.objects.filter(shift='3')
        releases = Release.objects.all().count()
        
        fallasr = Failures.objects.filter(sn_f__pn_b__commodity='RACK').count()
        fallass = Failures.objects.filter(sn_f__pn_b__commodity='SLED').count()
        fallask = Failures.objects.filter(sn_f__pn_b__commodity='KURA').count()
        turno1 = Release.objects.filter(shift='1').count()
        turno2 = Release.objects.filter(shift='2').count()
        turno3 = Release.objects.filter(shift='3').count()
        
        causeD = int(Failures.objects.filter(rootCause='DEBUG').count())
        causeR = int(Failures.objects.filter(rootCause='RUTEO').count())
        causeP = int(Failures.objects.filter(rootCause='PRUEBAS').count())
        causeE = int(Failures.objects.filter(rootCause='ENSAMBLE').count())
        causeC = int(Failures.objects.filter(rootCause='CODIGO').count())
        causeF = int(Failures.objects.filter(rootCause='FUNCIONAL').count())
        
        t1 = int(Failures.objects.filter(id_s__stationName='INIT').count())
        t2 = int(Failures.objects.filter(id_s__stationName='FVT_POWERSHELF').count())
        t3 = int(Failures.objects.filter(id_s__stationName='PRE_FVT_COMPUTE_SLED').count())
        t4 = int(Failures.objects.filter(id_s__stationName='FVT_COMPUTE_SLED').count())
        t5 = int(Failures.objects.filter(id_s__stationName='STRESS (SLED)').count())
        t6 = int(Failures.objects.filter(id_s__stationName='STRESS (KURA)').count())
        t7 = int(Failures.objects.filter(id_s__stationName='FVT_RACK').count())
        t8 = int(Failures.objects.filter(id_s__stationName='FVT_STORAGE').count())
        t9 = int(Failures.objects.filter(id_s__stationName='FVT_COMPUTE_MODULE').count())
    else:
        datetime.strptime(start, '%Y-%m-%d').date()
        datetime.strptime(end, '%Y-%m-%d').date()
        l1 = Release.objects.filter(shift='1').filter(date__gte=start, date__lt=end)
        l2 = Release.objects.filter(shift='2').filter(date__gte=start, date__lt=end)
        l3 = Release.objects.filter(shift='3').filter(date__gte=start, date__lt=end)
        releases = Release.objects.filter(date__gte=start, date__lt=end).count()
        
        turno1 = Release.objects.filter(shift='1').filter(date__gte=start, date__lt=end).count()
        turno2 = Release.objects.filter(shift='2').filter(date__gte=start, date__lt=end).count()
        turno3 = Release.objects.filter(shift='3').filter(date__gte=start, date__lt=end).count()
        
        fallasr = Failures.objects.filter(sn_f__pn_b__commodity='Rack').filter(failureDate__gte=start, failureDate__lt=end).count()
        fallass = Failures.objects.filter(sn_f__pn_b__commodity='Sled').filter(failureDate__gte=start, failureDate__lt=end).count()
        fallask = Failures.objects.filter(sn_f__pn_b__commodity='Kura').filter(failureDate__gte=start, failureDate__lt=end).count()

        causeD = int(Failures.objects.filter(rootCause='DEBUG').filter(failureDate__gte=start, failureDate__lt=end).count())
        causeR = int(Failures.objects.filter(rootCause='RUTEO').filter(failureDate__gte=start, failureDate__lt=end).count())
        causeP = int(Failures.objects.filter(rootCause='PRUEBAS').filter(failureDate__gte=start, failureDate__lt=end).count())
        causeE = int(Failures.objects.filter(rootCause='ENSAMBLE').filter(failureDate__gte=start, failureDate__lt=end).count())
        causeC = int(Failures.objects.filter(rootCause='CODIGO').filter(failureDate__gte=start, failureDate__lt=end).count())
        causeF = int(Failures.objects.filter(rootCause='FUNCIONAL').filter(failureDate__gte=start, failureDate__lt=end).count())
        
        t1 = int(Failures.objects.filter(id_s__stationName='INIT').filter(failureDate__gte=start, failureDate__lt=end).count())
        t2 = int(Failures.objects.filter(id_s__stationName='FVT_POWERSHELF').filter(failureDate__gte=start, failureDate__lt=end).count())
        t3 = int(Failures.objects.filter(id_s__stationName='PRE_FVT_COMPUTE_SLED').filter(failureDate__gte=start, failureDate__lt=end).count())
        t4 = int(Failures.objects.filter(id_s__stationName='FVT_COMPUTE_SLED').filter(failureDate__gte=start, failureDate__lt=end).count())
        t5 = int(Failures.objects.filter(id_s__stationName='STRESS (SLED)').filter(failureDate__gte=start, failureDate__lt=end).count())
        t6 = int(Failures.objects.filter(id_s__stationName='STRESS (KURA)').filter(failureDate__gte=start, failureDate__lt=end).count())
        t7 = int(Failures.objects.filter(id_s__stationName='FVT_RACK').filter(failureDate__gte=start, failureDate__lt=end).count())
        t8 = int(Failures.objects.filter(id_s__stationName='FVT_STORAGE').filter(failureDate__gte=start, failureDate__lt=end).count())
        t9 = int(Failures.objects.filter(id_s__stationName='FVT_COMPUTE_MODULE').filter(failureDate__gte=start, failureDate__lt=end).count())
    
    labels_test = ['INIT', 'FVT_POWERSHELF', 'PRE_FVT_COMPUTE_SLED', 'FVT_COMPUTE_SLED', 'STRESS (SLED)', 'STRESS (KURA)', 'FVT_RACK', 'FVT_STORAGE', 'FVT_COMPUTE_MODULE']
    value_test = [t1,t2,t3,t4,t5,t6,t7,t8,t9]
    
    labels_cause = ['Debug', 'Ruteo', 'Pruebas', 'Ensamble', 'Codigo', 'Funcional']
    value_cause = [causeD, causeR, causeP, causeE, causeC, causeF]
    
    labels_rack = ['Rack liberados','Rack fallados']
    value_rack = [int(releases),int(fallasr)]  
    
    labels_sled = ['Sled liberados','Sled fallados']
    value_sled = [(int(releases)*12),int(fallass)]
    
    labels_kura = ['Kura liberados','Kura fallados']
    value_kura = [int(releases),int(fallask)]
    
    turnos = [int(turno1), int(turno2), int(turno3)]
    seriales = [l1, l2, l3]
    
    context = {
        'turnos':turnos,
        'employe':employe, 
        'labels_rack':labels_rack, 
        'value_rack':value_rack, 
        'labels_sled':labels_sled, 
        'value_sled':value_sled, 
        'labels_kura':labels_kura, 
        'value_kura':value_kura, 
        'labels_cause':labels_cause, 
        'value_cause':value_cause, 
        'labels_test':labels_test, 
        'value_test':value_test, 
        'seriales':seriales,
        'is_supervisor': is_supervisor,
        'base_project': base_project
    }
    return render(request=request, template_name='base/table_release.html', context=context)

@login_required(login_url='login')
def manual_failure_registration(request):

    employee = Employes.objects.get(employeeNumber=request.user)

    is_supervisor = employee.privileges.endswith('S')
    base_project = employee.privileges[:-1] if is_supervisor else employee.privileges

    if 'bt-project' in request.POST: 
        if request.method == 'POST':
            employee.privileges = request.POST.get('bt-project')
            employee.save()
            return redirect('manual_failure_registration')
    
    if base_project == 'NA':
        return redirect('home')
    
    if request.method == 'POST':
        form = ManualFailureRegistrationForm(request.POST, request.FILES, project=base_project)
        if form.is_valid():
            try:
                current_hour = datetime.now().hour
                if 7 <= current_hour < 15:
                    shift = '1'
                elif 15 <= current_hour < 22.5:
                    shift = '2'
                else:
                    shift = '3'
                
                with transaction.atomic():
                    sn = form.cleaned_data['sn']
                    try:
                        uut = Uut.objects.get(sn=sn)
                        messages.info(request, f"UUT {sn} exists. Registering new failure.")
                    except Uut.DoesNotExist:
                        uut = Uut.objects.create(
                            sn=sn,
                            pn_b=form.cleaned_data['pn_b'],
                            employee_e=employee,
                            status=False
                        )
                        messages.success(request, f"New UUT {sn} created with failure record.")
                    
                    failure_status = form.cleaned_data['open_to_debug']
                    
                    failure = Failures.objects.create(
                        id_s=form.cleaned_data['id_s'],
                        sn_f=uut,
                        id_er=form.cleaned_data['id_er'],
                        shiftFailure=shift,
                        defectSymptom=form.cleaned_data['defectSymptom'],
                        analysis=form.cleaned_data['analysis'],
                        rootCauseCategory=form.cleaned_data['rootCauseCategory'],
                        correctiveActions=form.cleaned_data['correctiveActions'],
                        comments=form.cleaned_data['comments'],
                        employee_e=employee,
                        status=failure_status,
                        imgEvindence=form.cleaned_data['imgEvindence'],
                        log=form.cleaned_data['log']
                    )
                    
                    TestHistory.objects.create(
                        uut=uut,
                        station=failure.id_s,
                        employee_e=employee,
                        status=False,
                        test_date=timezone.now()
                    )
                
                messages.success(request, "Failure record created successfully!")
                return redirect('menuPruebas')
                
            except Exception as e:
                messages.error(request, f'Error registering failure: {str(e)}')
    else:
        form = ManualFailureRegistrationForm(project=base_project)
    
    context = {
        'form': form,
        'employee': employee,
        'title': 'Manual Failure Registration',
        'is_supervisor': is_supervisor,
        'base_project': base_project
    }
    return render(request, 'base/manual_failure_registration.html', context)

@login_required
def weekly_failure_report(request):
    try:
        employe = Employes.objects.get(employeeNumber=request.user)
    except Employes.DoesNotExist:
        return HttpResponse("Usuario no tiene perfil de empleado", status=403)
    
    # Manejar solicitud de descarga Excel primero
    if 'download_excel' in request.GET:
        return generate_excel_report(request, employe)
    
    # Determinar si es supervisor
    is_supervisor = employe.privileges.endswith('S')

    if not is_supervisor and not employe.QA:
        HttpResponse("Acceso denegado. Solo personal autorizado puede ver este reporte.", status=403)
        return redirect("menuMetricas")
    base_project = employe.privileges[:-1] if is_supervisor else employe.privileges
    
    # Obtener proyectos disponibles
    available_projects = get_available_projects(employe)
    
    # Procesar selección de proyecto
    selected_project = request.GET.get('project', base_project)
    if selected_project not in available_projects:
        selected_project = base_project
    
    # Procesar tipo de reporte
    report_type = request.GET.get('report_type', 'week')
    date_range = get_date_range(report_type, request)
    
    # Generar datos del reporte
    report_data = generate_report_data(selected_project, date_range['start'], date_range['end'], report_type)
    
    # Gráficas
    charts = create_charts(report_data, report_type)
    
    # Contexto para la plantilla
    context = {
        'employe': employe,
        'is_supervisor': is_supervisor,
        'available_projects': available_projects,
        'selected_project': selected_project,
        'report_type': report_type,
        'start_date': date_range['start'].strftime('%Y-%m-%d'),
        'end_date': date_range['end'].strftime('%Y-%m-%d'),
        'report_data': report_data,
        'charts': charts,
    }
    
    return render(request, 'base/report_form.html', context)

def generate_excel_report(request, employe):
    """Genera y descarga el reporte en formato Excel"""
    # Obtener parámetros del reporte
    project = request.GET.get('project', employe.privileges[:-1] if employe.privileges.endswith('S') else employe.privileges)
    report_type = request.GET.get('report_type', 'week')
    start_date = request.GET.get('start_date', None)
    end_date = request.GET.get('end_date', None)
    
    # Determinar rango de fechas
    date_range = get_date_range(report_type, request)
    
    # Obtener datos para Excel
    report_data = generate_report_data(project, date_range['start'], date_range['end'], report_type)
    
    # Crear DataFrame principal
    main_data = {
        'Período': [f"{date_range['start'].strftime('%Y-%m-%d')} a {date_range['end'].strftime('%Y-%m-%d')}"],
        'Proyecto': [project],
        'Total Pruebas': [report_data['total_tests']],
        'Pruebas Pasadas': [report_data['passed']],
        'Pruebas Falladas': [report_data['failed']],
        'Yield (%)': [report_data['yield_pct']],
        'Fallas Reales': [report_data['real_failures']],
        '% Fallas Reales': [report_data['real_failure_pct']],
        'NDF': [report_data['ndf_count']],
        '% NDF': [report_data['ndf_pct']],
    }
    df_main = pd.DataFrame(main_data)
    
    # Crear DataFrame de categorías de falla
    failure_categories = {
        'Categoría': ['Material', 'Workmanship', 'Operador', 'NDF'],
        'Cantidad': [
            report_data['material_count'],
            report_data['workmanship_count'],
            report_data['operator_count'],
            report_data['ndf_count']
        ]
    }
    df_failures = pd.DataFrame(failure_categories)
    
    # Crear DataFrame de estaciones
    station_data = {
        'Estación': list(report_data['station_data'].keys()),
        'Pruebas Realizadas': list(report_data['station_data'].values())
    }
    df_stations = pd.DataFrame(station_data)
    
    # Crear respuesta Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Reporte_Fallas_{project}_{date_range['start'].strftime('%Y%m%d')}_{date_range['end'].strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        # Hoja de resumen
        df_main.to_excel(writer, sheet_name='Resumen', index=False)
        
        # Hoja de categorías de falla
        df_failures.to_excel(writer, sheet_name='Categorías Falla', index=False)
        
        # Hoja de estaciones
        df_stations.to_excel(writer, sheet_name='Pruebas por Estación', index=False)
        
        # Ajustar anchos de columnas
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    return response

def get_available_projects(employe):
    """Devuelve los proyectos disponibles para el usuario basado en sus permisos"""
    projects = []
    if employe.pmd: projects.append('PMDU')
    if employe.dell: projects.append('DELL')
    if employe.switch: projects.append('1G-SW')
    if employe.sony: projects.append('SONY')
    return projects

def get_date_range(report_type, request):
    now = timezone.now()
    today = now.date()
    
    if report_type == 'day':
        # Definimos el rango de 7am a 5am del día siguiente
        if now.time() < time(5, 0):
            # Estamos en la ventana de 00:00-05:00, pertenece al día anterior (7am-5am)
            start_date = datetime.combine(today - timedelta(days=1), time(7, 0))
            end_date = datetime.combine(today, time(5, 0))
        elif now.time() < time(7, 0):
            # Estamos en la ventana de 05:00-07:00, día actual no ha comenzado
            start_date = datetime.combine(today, time(7, 0))
            end_date = datetime.combine(today + timedelta(days=1), time(5, 0))
        else:
            # Después de las 7am, día normal
            start_date = datetime.combine(today, time(7, 0))
            end_date = datetime.combine(today + timedelta(days=1), time(5, 0))
    elif report_type == 'week':
        start_date = datetime.combine(today - timedelta(days=today.weekday()), time(0, 0))
        end_date = datetime.combine(start_date.date() + timedelta(days=6), time(23, 59, 59))
    elif report_type == 'month':
        start_date = datetime.combine(today.replace(day=1), time(0, 0))
        end_date = datetime.combine((start_date.date() + timedelta(days=32)).replace(day=1) - timedelta(days=1), time(23, 59, 59))
    elif report_type == 'year':
        start_date = datetime.combine(today.replace(month=1, day=1), time(0, 0))
        end_date = datetime.combine(today.replace(month=12, day=31), time(23, 59, 59))
    else:  # custom
        start_date = request.GET.get('start_date', today)
        end_date = request.GET.get('end_date', today)
        if isinstance(start_date, str):
            start_date = datetime.strptime(start_date, '%Y-%m-%d')
        if isinstance(end_date, str):
            end_date = datetime.strptime(end_date, '%Y-%m-%d')
        # Para rangos personalizados, asumimos todo el día
        start_date = datetime.combine(start_date.date(), time(0, 0))
        end_date = datetime.combine(end_date.date(), time(23, 59, 59))
    
    return {'start': start_date, 'end': end_date}

def generate_report_data(project, start_date, end_date, report_type):
    # 1. Obtener pruebas dentro del período
    test_history = TestHistory.objects.filter(
        uut__pn_b__project=project,
        test_date__gte=start_date,
        test_date__lt=end_date
    ).select_related('uut', 'station').order_by('uut__sn', 'test_date')
    
    # 2. Inicializar variables
    total_tests = test_history.count()
    passed = test_history.filter(status=True).count()
    failed = test_history.filter(status=False).count()
    
    # Variables para categorías de falla (solo primera falla por SN)
    processed_sns = set()
    ndf_count = 0
    material_count = 0
    workmanship_count = 0
    operator_count = 0
    
    # Variables para reparaciones
    repaired_count = 0
    failed_sns_count = 0
    
    # Diccionarios para tracking de SNs por categoría
    material_sns = set()
    workmanship_sns = set()
    operator_sns = set()
    ndf_sns = set()
    
    # 3. Procesar cada SN único
    current_sn = None
    has_failed_in_period = False
    first_fail_date = None
    
    # Lista para trackear SNs que fallaron
    failed_sns_info = []
    
    for test in test_history:
        if test.uut.sn != current_sn:
            # Nuevo SN, procesar el anterior si falló
            if current_sn and has_failed_in_period:
                failed_sns_count += 1
                failed_sns_info.append({
                    'sn': current_sn,
                    'first_fail_date': first_fail_date
                })
            
            # Reiniciar para nuevo SN
            current_sn = test.uut.sn
            has_failed_in_period = False
            first_fail_date = None
        
        # Si es una falla dentro del período y es la primera falla de este SN
        if (not test.status and 
            (test.test_date >= start_date and test.test_date < end_date) and
            current_sn not in processed_sns):
            
            has_failed_in_period = True
            processed_sns.add(current_sn)
            
            if first_fail_date is None:
                first_fail_date = test.test_date
            
            # Buscar falla correspondiente en Failures
            related_failure = Failures.objects.filter(
                sn_f__sn=current_sn,
                failureDate__gte=test.test_date
            ).order_by('failureDate').first()
            
            if related_failure:
                category = related_failure.rootCauseCategory
                
                if category == 'Material':
                    material_count += 1
                    material_sns.add(current_sn)
                elif category == 'Workmanship':
                    workmanship_count += 1
                    workmanship_sns.add(current_sn)
                elif category == 'Operador':
                    operator_count += 1
                    operator_sns.add(current_sn)
                    ndf_count += 1
                    ndf_sns.add(current_sn)
                else:  # NDF
                    ndf_count += 1
                    ndf_sns.add(current_sn)
    
    # Procesar el último SN si falló
    if current_sn and has_failed_in_period:
        failed_sns_count += 1
        failed_sns_info.append({
            'sn': current_sn,
            'first_fail_date': first_fail_date
        })
    
    # 4. Verificar reparaciones para cada SN que falló
    for sn_info in failed_sns_info:
        # Buscar el primer PASS después de la primera falla dentro del período
        first_pass_after_fail = TestHistory.objects.filter(
            uut__sn=sn_info['sn'],
            status=True,
            test_date__gt=sn_info['first_fail_date'],
            test_date__gte=start_date,
            test_date__lt=end_date
        ).order_by('test_date').first()
        
        if first_pass_after_fail:
            repaired_count += 1
    
    # 5. Calcular valores
    real_failures = material_count + workmanship_count
    yield_pct = round((passed / total_tests * 100), 2) if total_tests > 0 else 0
    failure_pct = round((failed / total_tests * 100), 2) if total_tests > 0 else 0
    ndf_pct = round((ndf_count / total_tests * 100), 2) if total_tests > 0 else 0
    real_failure_pct = round((real_failures / total_tests * 100), 2) if total_tests > 0 else 0
    repair_rate = round((repaired_count / failed_sns_count * 100), 2) if failed_sns_count > 0 else 0
    
    # 6. Pruebas por estación
    stations = Station.objects.filter(stationProject=project)
    station_data = {
        station.stationName: test_history.filter(station=station).count()
        for station in stations
        if test_history.filter(station=station).exists()
    }
    
    # 7. Retornar estructura
    return {
        # Campos principales
        'total_tests': total_tests,
        'passed': passed,
        'failed': failed,
        'yield_pct': yield_pct,
        'failure_pct': failure_pct,
        'ndf_count': ndf_count,
        'ndf_pct': ndf_pct,
        'real_failures': real_failures,
        'real_failure_pct': real_failure_pct,
        'material_count': material_count,
        'workmanship_count': workmanship_count,
        'operator_count': operator_count,
        'station_data': station_data,
        'start_date': start_date,
        'end_date': end_date,
        
        # Campos de reparación
        'repaired_count': repaired_count,
        'repair_rate': repair_rate,
        'recovered_count': repaired_count,
        'recovery_rate': repair_rate,
        'failed_sns_count': failed_sns_count,
        
        # SNs por categoría
        'material_sns': list(material_sns),
        'workmanship_sns': list(workmanship_sns),
        'operator_sns': list(operator_sns),
        'ndf_sns': list(ndf_sns)
    }

def create_charts(report_data, report_type):
    charts = {}
    
    # Gráfica 1: Resumen de pruebas (ahora con recuperados)
    fig1 = go.Figure()
    fig1.add_trace(go.Indicator(
        mode="number",
        value=report_data['yield_pct'],
        number={'suffix': '%'},
        title={'text': "Yield"},
        domain={'row': 0, 'column': 0}
    ))
    fig1.add_trace(go.Indicator(
        mode="number",
        value=report_data['passed'],
        title={'text': "PASS"},
        domain={'row': 0, 'column': 1}
    ))
    fig1.add_trace(go.Indicator(
        mode="number",
        value=report_data['failed'],
        title={'text': "FAIL"},
        domain={'row': 0, 'column': 2}
    ))
    fig1.add_trace(go.Indicator(
        mode="number",
        value=report_data['recovered_count'],
        number={'suffix': f" ({report_data['recovery_rate']}%)"},
        title={'text': "Rapair Fail"},
        domain={'row': 1, 'column': 1}
    ))
    fig1.update_layout(
        grid={'rows': 2, 'columns': 3, 'pattern': "independent"},
        title="Resumen General"
    )
    charts['summary'] = plot(fig1, output_type='div')
    
    # Gráfica 2: Distribución de fallas
    fig2 = go.Figure()
    fig2.add_trace(go.Pie(
        labels=['Pasadas', 'Falladas'],
        values=[report_data['passed'], report_data['failed']],
        hole=.4
    ))
    fig2.update_layout(title="Distribución de Pruebas")
    charts['test_distribution'] = plot(fig2, output_type='div')
    
    # Gráfica 3: Categorías de falla (Operador como NDF)
    fig3 = go.Figure()
    fig3.add_trace(go.Bar(
        x=['Material', 'Workmanship', 'NDF'],  # Eliminado Operador
        y=[
            report_data['material_count'],
            report_data['workmanship_count'],
            report_data['ndf_count']  # Incluye los operadores
        ]
    ))
    fig3.update_layout(title="Categorías de Falla")
    charts['failure_categories'] = plot(fig3, output_type='div')
    
    # Gráfica 4: Pruebas por estación (ya filtradas en generate_report_data)
    station_names = list(report_data['station_data'].keys())
    station_counts = list(report_data['station_data'].values())
    
    fig4 = go.Figure()
    fig4.add_trace(go.Bar(
        x=station_names,
        y=station_counts
    ))
    fig4.update_layout(title="Pruebas por Estación")
    charts['station_tests'] = plot(fig4, output_type='div')
    
    return charts


def get_query_string_excluding(request, params_to_exclude):
    """
    Construye un query string excluyendo parámetros específicos
    """
    params = []
    exclude_list = params_to_exclude.split(',') if params_to_exclude else []
    
    for key, value in request.GET.items():
        if key not in exclude_list and value:
            params.append(f"{key}={value}")
    
    return '&'.join(params)

@login_required(login_url='login')
def project_yield_dashboard(request):
    try:
        # Autenticación y permisos
        try:
            employe = Employes.objects.get(employeeNumber=request.user)
        except Employes.DoesNotExist:
            return HttpResponse("Usuario no tiene perfil de empleado", status=403)
        
        # Manejar descarga Excel
        if 'download_excel' in request.GET:
            return generate_yield_excel_report(request, employe)
        
        # Configuración básica
        is_supervisor = employe.privileges.endswith('S')
        available_projects = get_available_projects(employe)
        report_type = request.GET.get('report_type', 'week')
        date_range = get_date_range(report_type, request)
        
        # Parámetros GET
        selected_project = request.GET.get('project')
        selected_station = request.GET.get('station')
        show_ndf_only = request.GET.get('ndf_only', 'false') == 'true'
        
        # Generar datos
        dashboard_data = generate_dashboard_data(
            available_projects,
            date_range['start'],
            date_range['end'],
            selected_project,
            selected_station,
            show_ndf_only
        )
        
        # Verificar si hay datos
        has_data = any(data['total_tests'] > 0 for data in dashboard_data['all_projects'].values())
        
        # Construir contexto
        context = {
            'employe': employe,
            'is_supervisor': is_supervisor,
            'available_projects': available_projects,
            'selected_project': selected_project,
            'selected_station': selected_station,
            'show_ndf_only': show_ndf_only,
            'report_type': report_type,
            'start_date': date_range['start'].strftime('%Y-%m-%d'),
            'end_date': date_range['end'].strftime('%Y-%m-%d'),
            'has_data': has_data,
            'request': request
        }
        
        if has_data:
            context['dashboard_data'] = dashboard_data
            
            # Obtener el primer proyecto para el resumen
            if dashboard_data['all_projects']:
                first_project_key = next(iter(dashboard_data['all_projects']))
                context['first_project'] = dashboard_data['all_projects'][first_project_key]
            
            # Obtener datos del proyecto seleccionado
            if selected_project and selected_project in dashboard_data['all_projects']:
                context['project_data'] = dashboard_data['all_projects'][selected_project]
            
            # Generar gráficas
            try:
                context['charts'] = create_interactive_charts(
                    dashboard_data, 
                    report_type, 
                    selected_project, 
                    selected_station
                )
            except Exception as e:
                print(f"Error creating charts: {str(e)}")
                context['chart_error'] = str(e)
        
        return render(request, 'base/yield_dashboard.html', context)

    except Exception as e:
        print(f"Error in view: {str(e)}")
        return HttpResponse(f"Error interno del servidor: {str(e)}", status=500)


def generate_dashboard_data(projects, start_date, end_date, selected_project=None, selected_station=None, show_ndf_only=False):
    all_projects_data = {}
    current_time = timezone.now()
    for project in projects:
        try:
            # Obtener todas las pruebas en el período
            test_history = TestHistory.objects.filter(
                uut__pn_b__project=project,
                test_date__gte=start_date,
                test_date__lt=end_date
            ).select_related('uut', 'station')
            
            # Obtener solo las pruebas fallidas
            failed_tests = test_history.filter(status=False)
            
            # Inicializar estructuras para conteo
            error_messages = []
            category_counts = {
                'Material': 0,
                'Workmanship': 0,
                'NDF': 0,
                'Operador': 0
            }
            
            # Procesar cada prueba fallida
            for test in failed_tests:
                # Buscar análisis de falla con margen de 1 minuto antes
                failure = Failures.objects.filter(
                    sn_f__sn=test.uut.sn,
                    failureDate__range=(
                        test.test_date - timedelta(minutes=1),  # 1 minuto antes
                        current_time  # Hasta el momento actual
                    )
                ).order_by('-failureDate').first()  # Tomar el MÁS RECIENTE
                
                if not failure:
                    # Si no encontramos en el rango cercano, buscar cualquier análisis para este SN
                    failure = Failures.objects.filter(
                        sn_f__sn=test.uut.sn
                    ).order_by('failureDate').first()
                
                if failure:
                    error_msg = failure.id_er.message if failure.id_er else "Error no especificado"
                    category = failure.rootCauseCategory
                    
                    # Actualizar contadores
                    if category == 'Operador':
                        category_counts['NDF'] += 1
                        category_counts['Operador'] += 1
                    else:
                        category_counts[category] += 1
                    
                    error_messages.append({
                        'id_er__message': error_msg,
                        'count': 1,  # Cada falla cuenta como 1
                        'category': category,
                        'station': test.station.stationName if test.station else None
                    })
                else:
                    # Si no hay falla registrada, contar como NDF
                    category_counts['NDF'] += 1
                    error_messages.append({
                        'id_er__message': "Falla no analizada",
                        'count': 1,
                        'category': 'NDF',
                        'station': test.station.stationName if test.station else None
                    })
            
            # Agrupar mensajes idénticos
            grouped_errors = {}
            for error in error_messages:
                key = (error['id_er__message'], error['category'])
                if key not in grouped_errors:
                    grouped_errors[key] = error.copy()
                    grouped_errors[key]['stations'] = set()
                else:
                    grouped_errors[key]['count'] += 1
                
                if error['station']:
                    grouped_errors[key]['stations'].add(error['station'])
            
            # Formatear estaciones para visualización
            sorted_errors = []
            for error in grouped_errors.values():
                error['stations'] = ', '.join(error['stations']) if error['stations'] else 'Varias'
                sorted_errors.append(error)
            
            # Ordenar por frecuencia
            sorted_errors.sort(key=lambda x: x['count'], reverse=True)
            
            # Cálculos básicos
            total_tests = test_history.count()
            passed = test_history.filter(status=True).count()
            failed = failed_tests.count()
            
            # Porcentajes
            yield_pct = round((passed / total_tests * 100), 2) if total_tests > 0 else 0
            failure_pct = round((failed / total_tests * 100), 2) if total_tests > 0 else 0
            ndf_pct = round((category_counts['NDF'] / total_tests * 100), 2) if total_tests > 0 else 0
            real_failure_pct = round(
                (category_counts['Material'] + category_counts['Workmanship']) / total_tests * 100, 2
            ) if total_tests > 0 else 0

            # Pruebas por estación
            station_data = {}
            stations = Station.objects.filter(stationProject=project)
            
            for station in stations:
                station_tests = test_history.filter(station=station)
                station_total = station_tests.count()
                
                if station_total > 0:
                    station_data[station.stationName] = {
                        'total': station_total,
                        'yield': round((station_tests.filter(status=True).count() / station_total * 100), 2)
                    }

            # Construir datos del proyecto
            project_data = {
                'total_tests': total_tests,
                'passed': passed,
                'failed': failed,
                'yield_pct': yield_pct,
                'failure_pct': failure_pct,
                'ndf_count': category_counts['NDF'],
                'ndf_pct': ndf_pct,
                'real_failures': category_counts['Material'] + category_counts['Workmanship'],
                'real_failure_pct': real_failure_pct,
                'material_count': category_counts['Material'],
                'workmanship_count': category_counts['Workmanship'],
                'operator_count': category_counts['Operador'],
                'station_data': station_data,
                'error_messages': sorted_errors[:10],  # Top 10 mensajes
                'start_date': start_date,
                'end_date': end_date,
            }
            
            all_projects_data[project] = project_data

        except Exception as e:
            print(f"Error procesando proyecto {project}: {str(e)}")
            all_projects_data[project] = {
                'error': str(e),
                'total_tests': 0,
                'passed': 0,
                'failed': 0,
                'yield_pct': 0,
                'failure_pct': 0,
                'ndf_count': 0,
                'ndf_pct': 0,
                'real_failures': 0,
                'real_failure_pct': 0,
                'material_count': 0,
                'workmanship_count': 0,
                'operator_count': 0,
                'station_data': {},
                'error_messages': [],
                'start_date': start_date,
                'end_date': end_date,
            }
    
    return {
        'all_projects': all_projects_data,
        'selected_project': selected_project,
        'selected_station': selected_station,
        'show_ndf_only': show_ndf_only,
        'start_date': start_date,
        'end_date': end_date,
    }

def create_interactive_charts(dashboard_data, report_type, selected_project=None, selected_station=None):
    charts = {}
    
    # 1. Gráfica de yield por proyecto
    fig1 = go.Figure()
    projects_data = dashboard_data['all_projects']
    
    if any(data['total_tests'] > 0 for data in projects_data.values()):
        for project, data in projects_data.items():
            fig1.add_trace(go.Bar(
                x=[project],
                y=[data['yield_pct']],
                name=project,
                customdata=[project],
                hovertemplate="<b>%{customdata}</b><br>Yield: %{y:.2f}%<extra></extra>"
            ))
    else:
        fig1.add_annotation(text="No hay datos disponibles",
                          xref="paper", yref="paper",
                          x=0.5, y=0.5, showarrow=False)
    
    fig1.update_layout(
        title="Yield por Proyecto",
        xaxis_title="Proyecto",
        yaxis_title="Yield (%)",
        hovermode="closest",
        clickmode='event+select'
    )
    charts['projects_yield'] = plot(fig1, output_type='div', include_plotlyjs=False)
    
    # 2. Gráfica de yield por estación (solo si hay proyecto seleccionado)
    if selected_project and selected_project in projects_data:
        project_data = projects_data[selected_project]
        fig2 = go.Figure()
        
        if project_data['station_data']:
            for station, data in project_data['station_data'].items():
                # Obtener datos reales para el tooltip
                tests = TestHistory.objects.filter(
                    uut__pn_b__project=selected_project,
                    station__stationName=station,
                    test_date__gte=dashboard_data['start_date'],
                    test_date__lt=dashboard_data['end_date']
                )
                total = tests.count()
                passed = tests.filter(status=True).count()
                failed = total - passed
                
                fig2.add_trace(go.Bar(
                    x=[station],
                    y=[data['yield']],
                    name=station,
                    customdata=[station],
                    hovertemplate=(
                        "<b>%{customdata}</b><br>" +
                        "Yield: %{y:.2f}%<br>" +
                        f"Total: {total}<br>" +
                        f"Pasadas: {passed}<br>" +
                        f"Fallidas: {failed}" +
                        "<extra></extra>"
                    )
                ))
            
            fig2.update_layout(
                title=f"Yield por Estación - {selected_project}",
                xaxis_title="Estación",
                yaxis_title="Yield (%)",
                hovermode="closest",
                clickmode='event+select'
            )
            charts['project_stations'] = plot(fig2, output_type='div', include_plotlyjs=False)
    
    return charts


def generate_yield_excel_report(request, employe):
    # Obtener parámetros igual que en la vista principal
    available_projects = get_available_projects(employe)
    report_type = request.GET.get('report_type', 'week')
    date_range = get_date_range(report_type, request)
    selected_project = request.GET.get('project')
    selected_station = request.GET.get('station')
    show_ndf_only = request.GET.get('ndf_only', 'false') == 'true'
    
    # Generar los datos
    dashboard_data = generate_dashboard_data(
        available_projects,
        date_range['start'],
        date_range['end'],
        selected_project,
        selected_station,
        show_ndf_only
    )
    
    # Crear el libro de Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"yield_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb = Workbook()
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    
    # Hoja 1: Resumen General
    ws_summary = wb.active
    ws_summary.title = "Resumen General"
    
    # Encabezados
    headers = ["Proyecto", "Total Pruebas", "Pass", "Fail", "Yield (%)"]
    
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws_summary[f"{col_letter}1"] = header
        ws_summary[f"{col_letter}1"].font = header_font
        ws_summary[f"{col_letter}1"].fill = header_fill
        ws_summary[f"{col_letter}1"].alignment = header_alignment
        ws_summary[f"{col_letter}1"].border = thin_border
    
    # Datos
    for row_num, (project, data) in enumerate(dashboard_data['all_projects'].items(), 2):
        ws_summary[f"A{row_num}"] = project
        ws_summary[f"B{row_num}"] = data['total_tests']
        ws_summary[f"C{row_num}"] = data['passed']
        ws_summary[f"D{row_num}"] = data['failed']
        ws_summary[f"E{row_num}"] = data['yield_pct'] / 100  # Formato porcentaje
        
        # Formato de porcentaje para yield
        ws_summary[f"E{row_num}"].number_format = '0.00%'
    
    # Ajustar anchos de columna
    for col in ws_summary.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws_summary.column_dimensions[column].width = adjusted_width
    
    # Estilos mejorados
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    error_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    even_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    for project, data in dashboard_data['all_projects'].items():
        ws_project = wb.create_sheet(title=project[:31])
        
        # Encabezado principal
        ws_project.merge_cells('A1:E1')
        ws_project['A1'] = f"Reporte de Yield - {project}"
        ws_project['A1'].font = Font(bold=True, size=14)
        ws_project['A1'].alignment = Alignment(horizontal='center')
        
        # Yield por Estación
        ws_project.append(['Yield por Estación'])
        ws_project.merge_cells(f'A{ws_project.max_row}:E{ws_project.max_row}')
        
        headers = ["Estación", "Total Pruebas", "Pasadas", "Fallidas", "Yield (%)"]
        ws_project.append(headers)
        
        # Aplicar estilo a encabezados
        for col in range(1, 6):
            cell = ws_project.cell(row=3, column=col)
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Datos de estaciones
        row_num = 4
        for station, station_data in data['station_data'].items():
            # Obtener datos reales
            tests = TestHistory.objects.filter(
                uut__pn_b__project=project,
                station__stationName=station,
                test_date__gte=date_range['start'],
                test_date__lt=date_range['end']
            )
            total = tests.count()
            passed = tests.filter(status=True).count()
            failed = total - passed
            yield_pct = (passed / total) if total > 0 else 0
            
            ws_project.append([
                station,
                total,
                passed,
                failed,
                yield_pct
            ])
            
            # Formato de porcentaje
            ws_project[f'E{row_num}'].number_format = '0.00%'
            
            # Autoajustar ancho basado en contenido
            for col in range(1, 6):
                cell = ws_project.cell(row=row_num, column=col)
                max_length = len(str(cell.value)) + 2
                col_letter = get_column_letter(col)
                if ws_project.column_dimensions[col_letter].width < max_length:
                    ws_project.column_dimensions[col_letter].width = max_length * 1.2
            
            row_num += 1
        
        # Mensajes de Error Globales
        ws_project.append([])
        ws_project.append(['Mensajes de Error Globales'])
        ws_project.merge_cells(f'A{ws_project.max_row}:C{ws_project.max_row}')
        
        error_headers = ["Mensaje", "Cantidad", "Categoría"]
        ws_project.append(error_headers)
        
        # Estilo encabezados de error
        for col in range(1, 4):
            cell = ws_project.cell(row=ws_project.max_row, column=col)
            cell.fill = error_fill
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Datos de errores
        for error in data['error_messages']:
            ws_project.append([
                error.get('id_er__message', 'Sin mensaje'),
                error.get('count', 0),
                error.get('category', 'Desconocida')
            ])
        
        # Mensajes de Error por Estación
        for station in data['station_data']:
            errors = Failures.objects.filter(
                sn_f__pn_b__project=project,
                id_s__stationName=station,
                failureDate__gte=date_range['start'],
                failureDate__lt=date_range['end']
            )
            
            if dashboard_data['show_ndf_only']:
                errors = errors.filter(rootCauseCategory='NDF')
            
            station_errors = list(errors.values(
                'id_er__message',
                'rootCauseCategory'
            ).annotate(
                count=models.Count('id'),
                category=models.F('rootCauseCategory')
            ).order_by('-count')[:10])
            
            if station_errors:
                ws_project.append([])
                ws_project.append([f'Mensajes de Error - Estación: {station}'])
                ws_project.merge_cells(f'A{ws_project.max_row}:C{ws_project.max_row}')
                
                ws_project.append(error_headers)
                
                # Estilo encabezados
                for col in range(1, 4):
                    cell = ws_project.cell(row=ws_project.max_row, column=col)
                    cell.fill = error_fill
                    cell.font = Font(color="FFFFFF", bold=True)
                
                for error in station_errors:
                    ws_project.append([
                        error.get('id_er__message', 'Sin mensaje'),
                        error.get('count', 0),
                        error.get('category', 'Desconocida')
                    ])
    
    # Eliminar hoja vacía inicial
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    wb.save(response)
    return response
